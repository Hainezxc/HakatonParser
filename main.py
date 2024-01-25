import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
from openpyxl.styles import Alignment
import json
import traceback
import os

def save_state(state):
    with open("state.json", "w") as file:
        json.dump(state, file)

def load_state():
    try:
        with open("state.json", "r") as file:
            return json.load(file)
    except FileNotFoundError:
        return {"page_counter": 1, "index": 0}

def setup_workbook():
    if os.path.exists("parsed_data.xlsx"):
        workbook = openpyxl.load_workbook("parsed_data.xlsx")
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ['Name', 'URL', 'Description', 'Event', 'Winner']
        sheet.append(headers)
        column_widths = {'A': 30, 'B': 20, 'C': 50, 'D': 20, 'E': 20, 'I': 20}
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
    return workbook

def main():
    brows = webdriver.Chrome()
    workbook = setup_workbook()
    sheet = workbook.active

    try:
        state = load_state()
        page_counter, index = state["page_counter"], state["index"]

        while True:
            elements = brows.find_elements(By.XPATH, '//*[@class="block border-2 border-black rounded overflow-hidden relative"]')

            if index >= len(elements):
                try:
                    if page_counter >= 283:
                        break
                    else:
                        brows.get(f'https://ethglobal.com/showcase/page/{page_counter}')
                        index = 0
                        page_counter += 1
                        time.sleep(1)
                        continue

                except:
                    print('PageError')
                    break

            project_row_index = sheet.max_row + 1

            elements[index].click()
            time.sleep(2)

            try:
                name = brows.find_element(By.XPATH, '//*[@class="text-4xl lg:text-5xl max-w-2xl mb-4"]').text
                sheet[f'A{project_row_index}'] = name
            except:
                print('NameError')
                name = 'None'
                sheet[f'A{project_row_index}'] = name

            try:
                parsed_url = brows.find_element(By.XPATH, '//*[contains(text(),"Source Code")]')
                href = parsed_url.get_attribute('href')
                url = href if href != 'https://github.com/' else 'URL does not exist'
                sheet[f'B{project_row_index}'] = url
            except:
                url = 'URL not found'
                sheet[f'B{project_row_index}'] = url

            try:
                description = brows.find_element(By.XPATH, '/html/body/div[1]/div/div[2]/div/div[2]/div[1]/div[2]/div[1]/p[1]').text
                sheet[f'C{project_row_index}'] = description
            except:
                print('DescError')
                description = 'None'
                sheet[f'C{project_row_index}'] = description

            try:
                event = brows.find_element(By.XPATH, '//*[contains(@class,"inline-flex overflow")]').text
                sheet[f'D{project_row_index}'] = event
            except:
                print('EventError')
                event = 'None'
                sheet[f'D{project_row_index}'] = event

            parsed_winners = brows.find_elements(By.XPATH, '//*[@class="font-normal"]')
            if parsed_winners:
                winners_text = ""
                for winner in parsed_winners:
                    winners_text += winner.text + " "
                winners_text = winners_text.strip()
                sheet[f'I{project_row_index}'] = winners_text
            else:
                sheet[f'I{project_row_index}'] = "Nothing in Winner"

            brows.back()
            time.sleep(1)
            index += 1

            save_state({"page_counter": page_counter, "index": index})

    except Exception as e:
        print(f"Error: {e}")
        traceback.print_exc()
        workbook.save('parsed_data.xlsx')
        brows.quit()
        time.sleep(30)
        main()

    finally:
        workbook.save('parsed_data.xlsx')
        brows.close()
        brows.quit()

if __name__ == "__main__":
    main()